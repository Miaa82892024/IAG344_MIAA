# processor.py
# Lógica de negocio: operaciones sobre Excel

import re
from openpyxl import load_workbook


def clean_id(value):
    """
    Elimina caracteres no numéricos de un identificador.
    """
    if value is None:
        return ""
    return re.sub(r"\D", "", str(value))


def merge_name(name, lastname):
    """
    Une nombre y apellido en un solo texto.
    """
    if name is None:
        name = ""
    if lastname is None:
        lastname = ""
    return f"{name} {lastname}".strip()


def ejecutar_accion(path):
    """
    Procesa el archivo Excel:
    - Limpia el identificador (columna A → D)
    - Une nombre y apellido (columnas B y C → E)
    """
    wb = load_workbook(path)

    # Acceso a la hoja llamada "Datos"
    ws = wb["Datos"]

    # Recorrer todas las filas desde la fila 2
    for row in range(2, ws.max_row + 1):
        # Columna D: identificador limpio
        ws[f"D{row}"] = clean_id(ws[f"A{row}"].value)

        # Columna E: nombre completo
        ws[f"E{row}"] = merge_name(
            ws[f"B{row}"].value,
            ws[f"C{row}"].value
        )

    # Guardar los cambios en el mismo archivo
    wb.save(path)


def process_excel_safe(path):
    """
    Ejecuta el procesamiento del Excel de forma segura,
    controlando errores comunes.
    """
    try:
        ejecutar_accion(path)
        return True, "Archivo procesado correctamente"

    except PermissionError:
        return (
            False,
            "El archivo Excel está abierto.\n"
            "Por favor, ciérrelo e intente de nuevo."
        )

    except KeyError:
        return False, "Hoja 'Datos' no encontrada"

    except Exception as e:
        return False, f"Error inesperado: {str(e)}"